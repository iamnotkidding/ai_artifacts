"""
Excel Trend Analyzer — Temperature & Humidity
- 데이터 로드: win32com.client (Excel COM)
- 3가지 입력 포맷 지원 (config.json의 formats 섹션)
- 전체 시트 각각 독립 처리
- 온도 / 습도 각각 독립 min_rows, min_val (config.json)
- 결과: Temp_Trend / Humid_Trend 컬럼 추가 후 _trend.xlsx 저장
"""

import json
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import win32com.client
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


# ══════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════
CONFIG_FILE = "config.json"

DEFAULT_CONFIG = {
    "input_format": "format1",
    "formats": {
        "format1": {
            "description": "Time, temp, humidity (헤더 1행, 데이터 2행~)",
            "header_row": 1,
            "data_start_row": 2,
            "columns": {"time": "Time", "temp": "temp", "humid": "humidity"}
        },
        "format2": {
            "description": "Date, Time, temp, humid (헤더 1행, 데이터 2행~, Date: yyyy/mm/dd)",
            "header_row": 1,
            "data_start_row": 2,
            "columns": {"date": "Date", "time": "Time", "temp": "temp", "humid": "humid"}
        },
        "format3": {
            "description": "빈칸, 시각, TEMP, HUMI (헤더 1행, 데이터 2행~)",
            "header_row": 1,
            "data_start_row": 2,
            "columns": {"index": "", "time": "시각", "temp": "TEMP", "humid": "HUMI"}
        }
    },
    "temp":  {"min_rows": 3, "min_val": 0.0},
    "humid": {"min_rows": 3, "min_val": 0.0},
}


def load_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        cfg = {}
        cfg["input_format"] = raw.get("input_format", DEFAULT_CONFIG["input_format"])
        # formats: 기본값 위에 사용자 값 덮어쓰기
        cfg["formats"] = {}
        for k, v in DEFAULT_CONFIG["formats"].items():
            merged = json.loads(json.dumps(v))          # deep copy
            merged.update(raw.get("formats", {}).get(k, {}))
            cfg["formats"][k] = merged
        # 사용자 정의 추가 포맷
        for k, v in raw.get("formats", {}).items():
            if k not in cfg["formats"]:
                cfg["formats"][k] = v
        for sensor in ("temp", "humid"):
            base = DEFAULT_CONFIG[sensor].copy()
            base.update(raw.get(sensor, {}))
            cfg[sensor] = base
        return cfg
    return json.loads(json.dumps(DEFAULT_CONFIG))


def save_config(cfg: dict):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# ══════════════════════════════════════════════
# WIN32 데이터 로드
# ══════════════════════════════════════════════
def load_sheet_via_win32(filepath: str, sheet_name: str,
                         header_row: int, data_start_row: int,
                         temp_col_name: str, humid_col_name: str):
    """
    win32com으로 Excel을 열어 지정 시트에서 데이터를 읽는다.
    반환: (headers: list[str], temp_values: list[float], humid_values: list[float])
    """
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False

    try:
        wb = xl.Workbooks.Open(os.path.abspath(filepath))
        ws = None
        for sh in wb.Sheets:
            if sh.Name == sheet_name:
                ws = sh
                break
        if ws is None:
            raise ValueError(f"시트 '{sheet_name}' 를 찾을 수 없습니다.")

        # 헤더 읽기
        last_col = ws.UsedRange.Columns.Count
        headers = []
        for c in range(1, last_col + 1):
            v = ws.Cells(header_row, c).Value
            headers.append(str(v).strip() if v is not None else "")

        def col_index(name: str) -> int:
            """헤더명 → 1-based 컬럼 인덱스. 빈 문자열이면 -1."""
            name = name.strip()
            if name == "":
                return -1
            for idx, h in enumerate(headers):
                if h == name:
                    return idx + 1
            return -1

        t_col  = col_index(temp_col_name)
        h_col  = col_index(humid_col_name)

        last_row = ws.UsedRange.Rows.Count

        def read_col(col_idx: int) -> list:
            if col_idx < 1:
                return []
            vals = []
            for r in range(data_start_row, last_row + 1):
                v = ws.Cells(r, col_idx).Value
                try:
                    vals.append(float(v) if v is not None else 0.0)
                except (TypeError, ValueError):
                    vals.append(0.0)
            return vals

        temp_vals  = read_col(t_col)
        humid_vals = read_col(h_col)

        wb.Close(False)
        return headers, temp_vals, humid_vals

    finally:
        xl.Quit()


def get_sheet_names_via_win32(filepath: str) -> list:
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(os.path.abspath(filepath))
        names = [sh.Name for sh in wb.Sheets]
        wb.Close(False)
        return names
    finally:
        xl.Quit()


def get_sheet_headers_via_win32(filepath: str, sheet_name: str,
                                header_row: int) -> list:
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(os.path.abspath(filepath))
        ws = None
        for sh in wb.Sheets:
            if sh.Name == sheet_name:
                ws = sh; break
        if ws is None:
            return []
        last_col = ws.UsedRange.Columns.Count
        headers = []
        for c in range(1, last_col + 1):
            v = ws.Cells(header_row, c).Value
            headers.append(str(v).strip() if v is not None else "")
        wb.Close(False)
        return headers
    finally:
        xl.Quit()


# ══════════════════════════════════════════════
# 추세 분석
# ══════════════════════════════════════════════
def analyze_trends(values: list, min_rows: int, min_val: float) -> list:
    n = len(values)
    if n == 0:
        return []

    raw = [""] * n
    for i in range(1, n):
        if values[i] > values[i - 1]:   raw[i] = "UP"
        elif values[i] < values[i - 1]: raw[i] = "DOWN"

    for i in range(1, n):
        if raw[i] == "": raw[i] = raw[i - 1]

    segments, i = [], 0
    while i < n:
        d = raw[i]; j = i
        while j < n and raw[j] == d: j += 1
        segments.append((i, j - 1, d)); i = j

    low_set = {k for k, v in enumerate(values) if v <= min_val}

    def is_noise(si):
        s, e, _ = segments[si]
        return all(k in low_set for k in range(s, e + 1)) or (e - s + 1) < min_rows

    resolved = [None] * len(segments)
    for si in range(len(segments)):
        if not is_noise(si): resolved[si] = segments[si][2]

    last = ""
    for si in range(len(segments)):
        if resolved[si] is not None: last = resolved[si]
        else: resolved[si] = last

    nxt = ""
    for si in range(len(segments) - 1, -1, -1):
        if not is_noise(si): nxt = segments[si][2]
        elif resolved[si] == "": resolved[si] = nxt

    result = [""] * n
    for si, (s, e, _) in enumerate(segments):
        for k in range(s, e + 1): result[k] = resolved[si] or ""
    return result


# ══════════════════════════════════════════════
# 스타일
# ══════════════════════════════════════════════
HEADER_FILL = PatternFill("solid", start_color="2E4057")
HEADER_FONT = Font(bold=True, color="FFFFFF")
T_UP_FILL   = PatternFill("solid", start_color="C8E6C9")
T_DOWN_FILL = PatternFill("solid", start_color="FFCDD2")
T_UP_FONT   = Font(color="1B5E20", bold=True)
T_DOWN_FONT = Font(color="B71C1C", bold=True)
H_UP_FILL   = PatternFill("solid", start_color="B3E5FC")
H_DOWN_FILL = PatternFill("solid", start_color="FFE0B2")
H_UP_FONT   = Font(color="01579B", bold=True)
H_DOWN_FONT = Font(color="E65100", bold=True)
CENTER      = Alignment(horizontal="center")


def write_trend_column(ws_write, data_start_row: int, trends: list,
                       result_col_name: str,
                       up_fill, down_fill, up_font, down_font):
    """openpyxl ws_write 에 결과 컬럼 추가"""
    existing = [ws_write.cell(row=1, column=c).value
                for c in range(1, ws_write.max_column + 1)]
    col_idx = (existing.index(result_col_name) + 1
               if result_col_name in existing
               else ws_write.max_column + 1)

    hc = ws_write.cell(row=1, column=col_idx, value=result_col_name)
    hc.font = HEADER_FONT; hc.fill = HEADER_FILL; hc.alignment = CENTER

    for i, trend in enumerate(trends):
        cell = ws_write.cell(row=data_start_row + i, column=col_idx, value=trend)
        cell.alignment = CENTER
        if trend == "UP":
            cell.fill = up_fill; cell.font = up_font
        elif trend == "DOWN":
            cell.fill = down_fill; cell.font = down_font

    ws_write.column_dimensions[
        ws_write.cell(row=1, column=col_idx).column_letter].width = 14


# ══════════════════════════════════════════════
# 전체 파일 처리
# ══════════════════════════════════════════════
def process_all_sheets(filepath, sheet_cfg_map, config, status_cb=None):
    def log(msg):
        if status_cb: status_cb(msg)

    fmt_key = config.get("input_format", "format1")
    fmt     = config["formats"][fmt_key]
    header_row    = int(fmt.get("header_row", 1))
    data_start_row = int(fmt.get("data_start_row", 2))
    col_map       = fmt.get("columns", {})
    temp_col_name  = col_map.get("temp",  "")
    humid_col_name = col_map.get("humid", "")

    t_cfg = config["temp"]
    h_cfg = config["humid"]

    log(f"입력 포맷: [{fmt_key}] {fmt.get('description','')}")
    log(f"헤더 행={header_row}, 데이터 시작 행={data_start_row}")
    log(f"온도 컬럼='{temp_col_name}', 습도 컬럼='{humid_col_name}'")

    wb_write  = openpyxl.load_workbook(filepath)
    processed = []

    for sheet_name, scfg in sheet_cfg_map.items():
        if not scfg.get("enabled", True):
            log(f"  [{sheet_name}] 스킵 (비활성)")
            continue

        log(f"  [{sheet_name}] win32 로드 중...")
        try:
            headers, temp_vals, humid_vals = load_sheet_via_win32(
                filepath, sheet_name,
                header_row, data_start_row,
                temp_col_name, humid_col_name)
        except Exception as e:
            log(f"  [{sheet_name}] 로드 실패: {e}"); continue

        ws_w = wb_write[sheet_name]
        sheet_done = False

        if temp_vals:
            trends = analyze_trends(temp_vals,
                                    int(t_cfg["min_rows"]),
                                    float(t_cfg["min_val"]))
            write_trend_column(ws_w, data_start_row, trends, "Temp_Trend",
                               T_UP_FILL, T_DOWN_FILL, T_UP_FONT, T_DOWN_FONT)
            log(f"  [{sheet_name}] 🌡 온도({temp_col_name})  "
                f"min_rows={t_cfg['min_rows']}, min_val={t_cfg['min_val']}  "
                f"→ UP={trends.count('UP')}, DOWN={trends.count('DOWN')}")
            sheet_done = True
        else:
            log(f"  [{sheet_name}] 🌡 온도 데이터 없음 → 스킵")

        if humid_vals:
            trends = analyze_trends(humid_vals,
                                    int(h_cfg["min_rows"]),
                                    float(h_cfg["min_val"]))
            write_trend_column(ws_w, data_start_row, trends, "Humid_Trend",
                               H_UP_FILL, H_DOWN_FILL, H_UP_FONT, H_DOWN_FONT)
            log(f"  [{sheet_name}] 💧 습도({humid_col_name})  "
                f"min_rows={h_cfg['min_rows']}, min_val={h_cfg['min_val']}  "
                f"→ UP={trends.count('UP')}, DOWN={trends.count('DOWN')}")
            sheet_done = True
        else:
            log(f"  [{sheet_name}] 💧 습도 데이터 없음 → 스킵")

        if sheet_done:
            processed.append(sheet_name)

    base, ext = os.path.splitext(filepath)
    out_path = base + "_trend" + ext
    wb_write.save(out_path)
    log(f"\n저장: {out_path}")
    log(f"처리 완료 시트: {len(processed)}개")

    log("Excel에서 파일 여는 중...")
    try:
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = True
        xl.Workbooks.Open(os.path.abspath(out_path))
        log("Excel 열기 완료!")
    except Exception as e:
        log(f"[경고] Win32 Excel 열기 실패: {e}")

    return out_path, processed


# ══════════════════════════════════════════════
# GUI: 포맷 에디터 다이얼로그
# ══════════════════════════════════════════════
class FormatEditorDialog(tk.Toplevel):
    """단일 포맷의 header_row / data_start_row / columns 편집 다이얼로그"""

    def __init__(self, parent, fmt_key: str, fmt_data: dict, bg, fg, ebg, lf):
        super().__init__(parent)
        self.title(f"포맷 편집 — {fmt_key}")
        self.configure(bg=bg)
        self.resizable(False, False)
        self.result = None

        self._bg = bg; self._fg = fg; self._ebg = ebg; self._lf = lf
        self._vars = {}

        P = 10
        tk.Label(self, text=f"[{fmt_key}] 설정",
                 font=(lf[0], 11, "bold"), bg=bg, fg=fg).pack(pady=(P, 4), padx=P)

        frm = tk.Frame(self, bg=bg)
        frm.pack(padx=P, pady=4)

        def row_entry(label, key, val, r):
            tk.Label(frm, text=label, font=lf, bg=bg, fg=fg,
                     anchor="e", width=18).grid(row=r, column=0, padx=(0,6), pady=3)
            v = tk.StringVar(value=str(val))
            tk.Entry(frm, textvariable=v, width=22,
                     bg=ebg, fg=fg, insertbackground=fg,
                     relief="flat").grid(row=r, column=1, pady=3)
            self._vars[key] = v

        row_entry("description",   "description",    fmt_data.get("description", ""), 0)
        row_entry("header_row",    "header_row",      fmt_data.get("header_row", 1),   1)
        row_entry("data_start_row","data_start_row",  fmt_data.get("data_start_row",2), 2)

        tk.Frame(self, bg="#45475A", height=1).pack(fill="x", padx=P, pady=6)
        tk.Label(self, text="columns (헤더명 매핑)",
                 font=(lf[0], 10, "bold"), bg=bg, fg="#A6E3A1").pack()

        cols_frm = tk.Frame(self, bg=bg)
        cols_frm.pack(padx=P, pady=4)
        self._col_vars = {}
        col_defs = fmt_data.get("columns", {})
        # 표준 키 순서 보장
        col_keys = list(dict.fromkeys(
            ["time", "date", "temp", "humid", "index"] + list(col_defs.keys())))
        for ri, k in enumerate(col_keys):
            tk.Label(cols_frm, text=k, font=lf, bg=bg, fg=fg,
                     anchor="e", width=10).grid(row=ri, column=0, padx=(0,6), pady=2)
            v = tk.StringVar(value=col_defs.get(k, ""))
            tk.Entry(cols_frm, textvariable=v, width=22,
                     bg=ebg, fg=fg, insertbackground=fg,
                     relief="flat").grid(row=ri, column=1, pady=2)
            self._col_vars[k] = v

        tk.Frame(self, bg="#45475A", height=1).pack(fill="x", padx=P, pady=6)
        bf = tk.Frame(self, bg=bg)
        bf.pack(pady=(0, P))
        tk.Button(bf, text="저장", bg="#89B4FA", fg="#1E1E2E",
                  font=lf, relief="flat", cursor="hand2",
                  command=self._save).pack(side="left", ipadx=10, padx=4)
        tk.Button(bf, text="취소", bg="#45475A", fg=fg,
                  font=lf, relief="flat", cursor="hand2",
                  command=self.destroy).pack(side="left", ipadx=10, padx=4)

        self.grab_set()
        self.wait_window()

    def _save(self):
        try:
            self.result = {
                "description":    self._vars["description"].get(),
                "header_row":     int(self._vars["header_row"].get()),
                "data_start_row": int(self._vars["data_start_row"].get()),
                "columns": {k: v.get() for k, v in self._col_vars.items()
                            if v.get() != "" or k in ("time", "temp", "humid")}
            }
        except ValueError:
            messagebox.showerror("오류", "header_row / data_start_row 는 정수로 입력하세요.")
            return
        self.destroy()


# ══════════════════════════════════════════════
# GUI: 센서 설정 프레임
# ══════════════════════════════════════════════
class SensorConfigFrame(tk.LabelFrame):
    def __init__(self, parent, label, init: dict, bg, fg, entry_bg, lf):
        super().__init__(parent, text=label, font=(lf[0], 10, "bold"),
                         bg=bg, fg=fg, bd=1, relief="groove",
                         labelanchor="nw", padx=8, pady=6)
        self.configure(background=bg)
        self.min_rows_var = tk.StringVar(value=str(init.get("min_rows", 3)))
        self.min_val_var  = tk.StringVar(value=str(init.get("min_val",  0.0)))

        for i, (lbl, var) in enumerate([
            ("min_rows  (노이즈 최소 행)", self.min_rows_var),
            ("min_val   (노이즈 임계값)",  self.min_val_var),
        ]):
            tk.Label(self, text=lbl, font=lf, bg=bg, fg=fg).grid(
                row=i, column=0, sticky="e", padx=(0, 6), pady=3)
            tk.Entry(self, textvariable=var, width=10,
                     bg=entry_bg, fg=fg, insertbackground=fg,
                     relief="flat").grid(row=i, column=1, sticky="w", pady=3)

    def get(self) -> dict:
        return {"min_rows": int(self.min_rows_var.get()),
                "min_val":  float(self.min_val_var.get())}


# ══════════════════════════════════════════════
# GUI: 시트 1행
# ══════════════════════════════════════════════
class SheetRow:
    def __init__(self, parent, sheet_name, bg, fg, lf, row_idx):
        self.sheet_name = sheet_name
        self.enabled = tk.BooleanVar(value=True)

        tk.Checkbutton(parent, variable=self.enabled, bg=bg, fg=fg,
                       selectcolor="#313244", activebackground=bg).grid(
            row=row_idx, column=0, padx=(6, 2))
        tk.Label(parent, text=sheet_name, font=lf, bg=bg, fg=fg,
                 width=24, anchor="w").grid(row=row_idx, column=1, padx=6)

    def get(self) -> dict:
        return {"enabled": self.enabled.get()}


# ══════════════════════════════════════════════
# GUI: 메인 앱
# ══════════════════════════════════════════════
class App(tk.Tk):
    BG  = "#1E1E2E";  FG  = "#CDD6F4"
    EBG = "#313244";  BBG = "#89B4FA";  BFG = "#1E1E2E"
    SEP = "#45475A"
    LF  = ("Segoe UI", 10)
    TF  = ("Segoe UI", 13, "bold")
    PAD = 12

    def __init__(self):
        super().__init__()
        self.title("Trend Analyzer — Temperature & Humidity")
        self.configure(bg=self.BG)
        self.resizable(True, False)
        self.config_data = load_config()
        self._sheet_rows: list[SheetRow] = []
        self._build_ui()

    # ── UI ──────────────────────────────────────
    def _build_ui(self):
        P = self.PAD

        tk.Label(self, text="🌡💧 Temperature & Humidity Trend Analyzer",
                 font=self.TF, bg=self.BG, fg=self.BBG).pack(pady=(P, 4))

        # 파일 선택
        ff = tk.Frame(self, bg=self.BG)
        ff.pack(fill="x", padx=P, pady=4)
        tk.Label(ff, text="Excel 파일", font=self.LF,
                 bg=self.BG, fg=self.FG, width=10, anchor="e").pack(side="left")
        self.file_var = tk.StringVar()
        tk.Entry(ff, textvariable=self.file_var, bg=self.EBG, fg=self.FG,
                 insertbackground=self.FG, relief="flat").pack(
            side="left", fill="x", expand=True, padx=6)
        tk.Button(ff, text="찾아보기", bg=self.BBG, fg=self.BFG,
                  font=self.LF, relief="flat", cursor="hand2",
                  command=self._browse).pack(side="left")

        self._sep()

        # ── 입력 포맷 선택 ──────────────────────────
        tk.Label(self, text="📂 입력 포맷 설정",
                 font=("Segoe UI", 10, "bold"),
                 bg=self.BG, fg="#A6E3A1").pack(pady=(0, 4))

        fmt_outer = tk.Frame(self, bg=self.BG)
        fmt_outer.pack(fill="x", padx=P, pady=4)

        # 포맷 선택 콤보
        fsel = tk.Frame(fmt_outer, bg=self.BG)
        fsel.pack(fill="x")
        tk.Label(fsel, text="사용 포맷", font=self.LF,
                 bg=self.BG, fg=self.FG, width=10, anchor="e").pack(side="left")
        self.fmt_var = tk.StringVar(
            value=self.config_data.get("input_format", "format1"))
        self.fmt_cb = ttk.Combobox(fsel, textvariable=self.fmt_var,
                                   width=12, state="readonly")
        self.fmt_cb.pack(side="left", padx=6)
        self.fmt_cb.bind("<<ComboboxSelected>>", self._on_fmt_select)

        tk.Button(fsel, text="편집", bg="#CBA6F7", fg=self.BFG,
                  font=self.LF, relief="flat", cursor="hand2",
                  command=self._edit_format).pack(side="left", padx=4)

        # 포맷 요약 표시
        self.fmt_info_var = tk.StringVar()
        tk.Label(fmt_outer, textvariable=self.fmt_info_var,
                 font=("Consolas", 9), bg=self.BG, fg="#A6ADC8",
                 justify="left", anchor="w").pack(fill="x", pady=(4, 0))

        self._refresh_fmt_combo()
        self._refresh_fmt_info()

        self._sep()

        # ── 시트 목록 ───────────────────────────────
        tk.Label(self, text="📋 시트 활성화",
                 font=("Segoe UI", 10, "bold"),
                 bg=self.BG, fg="#A6E3A1").pack(pady=(0, 2))

        hf = tk.Frame(self, bg=self.BG)
        hf.pack(fill="x", padx=P)
        for col, txt, w in [(0,"적용",4),(1,"시트명",24)]:
            tk.Label(hf, text=txt, font=("Segoe UI", 9, "bold"),
                     bg="#313244", fg="#A6ADC8", width=w,
                     padx=4, pady=2).grid(row=0, column=col, padx=2, pady=(0,2))

        self.canvas = tk.Canvas(self, bg=self.BG, highlightthickness=0, height=130)
        vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        self.canvas.pack(side="left", fill="both", expand=True, padx=(P, 0))
        vsb.pack(side="left", fill="y", padx=(0, P))

        self.sf = tk.Frame(self.canvas, bg=self.BG)
        self._cw = self.canvas.create_window((0, 0), window=self.sf, anchor="nw")
        self.sf.bind("<Configure>",
                     lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>",
                         lambda e: self.canvas.itemconfig(self._cw, width=e.width))

        tk.Label(self.sf, text="← Excel 파일을 선택하면 시트 목록이 표시됩니다",
                 font=self.LF, bg=self.BG, fg="#585B70").grid(
            row=0, column=0, columnspan=2, pady=14)

        self._sep()

        # ── 센서 설정 ───────────────────────────────
        tk.Label(self, text="⚙ 설정 (config.json)",
                 font=("Segoe UI", 10, "bold"),
                 bg=self.BG, fg="#A6E3A1").pack(pady=(0, 4))

        cfg_row = tk.Frame(self, bg=self.BG)
        cfg_row.pack(padx=P, pady=4)
        self.temp_cfg_frame = SensorConfigFrame(
            cfg_row, "🌡 온도 (temp)",
            self.config_data["temp"],
            self.BG, self.FG, self.EBG, self.LF)
        self.temp_cfg_frame.grid(row=0, column=0, padx=(0, 12), sticky="n")

        self.humid_cfg_frame = SensorConfigFrame(
            cfg_row, "💧 습도 (humid)",
            self.config_data["humid"],
            self.BG, self.FG, self.EBG, self.LF)
        self.humid_cfg_frame.grid(row=0, column=1, sticky="n")

        tk.Button(cfg_row, text="설정\n저장", bg="#A6E3A1", fg=self.BFG,
                  font=self.LF, relief="flat", cursor="hand2",
                  command=self._save_cfg).grid(row=0, column=2, padx=(12,0),
                                               ipadx=6, ipady=4, sticky="ns")

        self._sep()

        tk.Button(self, text="▶  전체 시트 분석 실행",
                  bg=self.BBG, fg=self.BFG,
                  font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2",
                  command=self._run).pack(pady=(0, 6), ipadx=20, ipady=4)

        self.log_box = tk.Text(self, height=8, bg="#181825", fg="#BAC2DE",
                               font=("Consolas", 9), relief="flat", state="disabled")
        self.log_box.pack(fill="x", padx=P, pady=(0, P))

    def _sep(self):
        tk.Frame(self, bg=self.SEP, height=1).pack(fill="x", padx=self.PAD, pady=6)

    # ── 포맷 관련 ────────────────────────────────
    def _refresh_fmt_combo(self):
        keys = list(self.config_data["formats"].keys())
        self.fmt_cb["values"] = keys
        if self.fmt_var.get() not in keys and keys:
            self.fmt_var.set(keys[0])

    def _refresh_fmt_info(self):
        key = self.fmt_var.get()
        fmt = self.config_data["formats"].get(key, {})
        cols = fmt.get("columns", {})
        col_str = "  |  ".join(f"{k}='{v}'" for k, v in cols.items() if v)
        info = (f"  헤더 행: {fmt.get('header_row',1)}  "
                f"데이터 시작: {fmt.get('data_start_row',2)}행\n"
                f"  {col_str}")
        self.fmt_info_var.set(info)

    def _on_fmt_select(self, _=None):
        self._refresh_fmt_info()

    def _edit_format(self):
        key = self.fmt_var.get()
        fmt_data = self.config_data["formats"].get(key, {})
        dlg = FormatEditorDialog(self, key, fmt_data,
                                 self.BG, self.FG, self.EBG, self.LF)
        if dlg.result:
            self.config_data["formats"][key] = dlg.result
            self._refresh_fmt_info()
            self._log(f"포맷 [{key}] 수정됨")

    # ── 파일 로드 ─────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")])
        if not path: return
        self.file_var.set(path)
        self._load_sheets(path)

    def _load_sheets(self, path):
        for w in self.sf.winfo_children(): w.destroy()
        self._sheet_rows.clear()

        fmt_key = self.fmt_var.get()
        fmt     = self.config_data["formats"].get(fmt_key, {})
        hdr_row = int(fmt.get("header_row", 1))

        self._log(f"시트 목록 로드 중 (win32)...")
        self.update_idletasks()
        try:
            names = get_sheet_names_via_win32(path)
        except Exception as e:
            messagebox.showerror("오류", f"파일 읽기 실패:\n{e}"); return

        for i, sname in enumerate(names):
            sr = SheetRow(self.sf, sname,
                          self.BG, self.FG, self.LF, i)
            self._sheet_rows.append(sr)

        self._log(f"파일 로드 완료: {len(names)}개 시트 — " + ", ".join(names))

    # ── 설정 저장 ─────────────────────────────────
    def _save_cfg(self):
        try:
            cfg = self.config_data.copy()
            cfg["input_format"] = self.fmt_var.get()
            cfg["temp"]  = self.temp_cfg_frame.get()
            cfg["humid"] = self.humid_cfg_frame.get()
            save_config(cfg)
            self.config_data = cfg
            self._log(
                f"설정 저장 완료  포맷=[{cfg['input_format']}]\n"
                f"  🌡 temp : min_rows={cfg['temp']['min_rows']}, min_val={cfg['temp']['min_val']}\n"
                f"  💧 humid: min_rows={cfg['humid']['min_rows']}, min_val={cfg['humid']['min_val']}")
        except ValueError:
            messagebox.showerror("오류", "min_rows는 정수, min_val은 실수로 입력하세요.")

    # ── 분석 실행 ─────────────────────────────────
    def _run(self):
        filepath = self.file_var.get().strip()
        if not filepath or not os.path.exists(filepath):
            messagebox.showerror("오류", "유효한 Excel 파일을 선택하세요."); return
        if not self._sheet_rows:
            messagebox.showerror("오류", "시트 정보가 없습니다. 파일을 다시 선택하세요."); return

        try:
            cfg = self.config_data.copy()
            cfg["input_format"] = self.fmt_var.get()
            cfg["temp"]  = self.temp_cfg_frame.get()
            cfg["humid"] = self.humid_cfg_frame.get()
        except ValueError:
            messagebox.showerror("오류", "min_rows / min_val 값을 확인하세요."); return

        sheet_cfg_map = {sr.sheet_name: sr.get() for sr in self._sheet_rows}

        self._log("═" * 56)
        self._log(f"파일: {os.path.basename(filepath)}")

        def status(msg):
            self._log(msg); self.update_idletasks()

        try:
            out_path, done = process_all_sheets(
                filepath, sheet_cfg_map, cfg, status_cb=status)
            messagebox.showinfo(
                "완료",
                f"분석 완료!\n처리 시트: {len(done)}개\n저장: {os.path.basename(out_path)}")
        except Exception as e:
            self._log(f"[오류] {e}"); messagebox.showerror("오류", str(e))

    def _log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")


# ══════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()
